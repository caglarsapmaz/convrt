import openpyxl
import csv

def convert_xlsx_to_csv(input_path: str, output_path: str):
    """Excel (XLSX) dosyasındaki verileri CSV formatına dönüştürür."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    sheet = wb.active
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

def convert_csv_to_xlsx(input_path: str, output_path: str):
    """CSV dosyasındaki verileri Excel (XLSX) formatına dönüştürür."""
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(input_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            sheet.append(row)
    wb.save(output_path)